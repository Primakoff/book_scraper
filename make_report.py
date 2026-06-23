# -*- coding: utf-8 -*-
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", start_color="2F5496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BASE_FONT = Font(name="Arial", size=11)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("title", "Title", 55, None),
    ("price", "Price (£)", 12, "#,##0.00"),
    ("rating", "Rating", 10, "0"),
    ("availability", "Availability", 16, None),
    ("url", "URL", 50, None),
]

INPUT = Path("data/books.json")
OUTPUT = Path("data/books_report.xlsx")


def main() -> None:
    data = json.loads(INPUT.read_text(encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"

    for col, (_, header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    for row, item in enumerate(data, start=2):
        for col, (key, _, _, fmt) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row, column=col, value=item.get(key, ""))
            cell.font = BASE_FONT
            cell.border = BORDER
            if fmt:
                cell.number_format = fmt

    last = len(data) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last}"
    ws.freeze_panes = "A2"

    s = wb.create_sheet("Сводка")
    s.column_dimensions["A"].width = 24
    s.column_dimensions["B"].width = 14
    s.cell(row=1, column=1, value="Сводка по каталогу").font = Font(
        name="Arial", bold=True, size=14)
    rows = [
        ("Всего книг", f"=COUNTA('Каталог'!A2:A{last})", "0"),
        ("Средняя цена (£)", f"=AVERAGE('Каталог'!B2:B{last})", "#,##0.00"),
        ("Мин. цена (£)", f"=MIN('Каталог'!B2:B{last})", "#,##0.00"),
        ("Макс. цена (£)", f"=MAX('Каталог'!B2:B{last})", "#,##0.00"),
        ("Средний рейтинг", f"=AVERAGE('Каталог'!C2:C{last})", "0.0"),
    ]
    r = 3
    for label, formula, fmt in rows:
        s.cell(row=r, column=1, value=label).font = BASE_FONT
        c = s.cell(row=r, column=2, value=formula)
        c.font = BASE_FONT
        c.number_format = fmt
        r += 1

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Готово: {OUTPUT} ({len(data)} книг)")


if __name__ == "__main__":
    main()